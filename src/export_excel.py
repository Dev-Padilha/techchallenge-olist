"""
export_excel.py — Gera o dashboard executivo em Excel (.xlsx) a partir de data/bi/.

Um único arquivo `dashboard/dashboard_olist.xlsx` com uma página "Dashboard" que conta a
mesma história do relatorio (Crescimento -> Risco logistico -> Atraso destroi reputacao ->
Valor de agir), alimentada por abas de dados nativas do Excel.

Uso:
    python src/export_excel.py
Pre-requisito: data/bi/*.csv (gere com `python src/export_bi.py`).
"""
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

ROOT = Path(__file__).resolve().parents[1]
BI = ROOT / "data" / "bi"
OUT = ROOT / "dashboard" / "dashboard_olist.xlsx"

# Paleta executiva (mesma das figuras/relatorio)
AZUL = "1F4E79"
VERMELHO = "C0392B"
VERDE = "2E8B57"
LARANJA = "E07B39"
CINZA = "7F8C8D"
CINZA_CLARO = "F2F4F5"
BRANCO = "FFFFFF"

thin = Side(style="thin", color="D9D9D9")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def load(name):
    return pd.read_csv(BI / f"{name}.csv")


def write_data_sheet(wb, df, title):
    """Cria uma aba de dados (fonte dos graficos)."""
    ws = wb.create_sheet(title)
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    # cabecalho em negrito
    for c in ws[1]:
        c.font = Font(bold=True, color=BRANCO)
        c.fill = PatternFill("solid", fgColor=AZUL)
    ws.sheet_state = "visible"
    return ws


def kpi_card(ws, anchor_col, row, titulo, valor, cor=AZUL):
    """Desenha um cartao de KPI ocupando 3 colunas x 3 linhas a partir de (row, anchor_col)."""
    c0 = anchor_col
    ws.merge_cells(start_row=row, start_column=c0, end_row=row, end_column=c0 + 2)
    ws.merge_cells(start_row=row + 1, start_column=c0, end_row=row + 2, end_column=c0 + 2)
    t = ws.cell(row=row, column=c0, value=titulo)
    t.font = Font(bold=True, size=10, color=BRANCO)
    t.alignment = Alignment(horizontal="center", vertical="center")
    v = ws.cell(row=row + 1, column=c0, value=valor)
    v.font = Font(bold=True, size=20, color=cor)
    v.alignment = Alignment(horizontal="center", vertical="center")
    for rr in range(row, row + 3):
        for cc in range(c0, c0 + 3):
            cell = ws.cell(row=rr, column=cc)
            cell.border = BORDER
            if rr == row:
                cell.fill = PatternFill("solid", fgColor=cor)
            else:
                cell.fill = PatternFill("solid", fgColor=BRANCO)


def value_labels():
    """Rotulo de dados mostrando SO o valor (sem nome da serie nem categoria)."""
    dl = DataLabelList()
    dl.showVal = True
    dl.showSerName = False
    dl.showCatName = False
    dl.showLegendKey = False
    dl.showPercent = False
    dl.showBubbleSize = False
    return dl


def style_chart(chart):
    """Tira as linhas de grade e deixa o grafico com cara de cartao de dashboard."""
    for ax in (chart.y_axis, chart.x_axis):
        ax.majorGridlines = None
        ax.minorGridlines = None
        ax.delete = False
    # fundo branco do grafico, borda fina (cartao) sobre o canvas cinza
    chart.graphical_properties = GraphicalProperties(
        solidFill="FFFFFF", ln=LineProperties(solidFill="E2E6EA", w=9525))
    # area de plotagem sem preenchimento (so o branco do cartao aparece)
    chart.plot_area.graphicalProperties = GraphicalProperties(noFill=True)


def main():
    kpis = load("kpis")
    mensal = load("agg_mensal")
    atraso = load("tab_atraso_nota")
    simul = load("tab_simulacao")
    categoria = load("agg_categoria").sort_values("receita", ascending=False).head(8)
    estado = load("agg_estado").sort_values("pedidos", ascending=False).head(12)

    wb = Workbook()
    dash = wb.active
    dash.title = "Dashboard"
    dash.sheet_view.showGridLines = False

    # ---- abas de dados ----
    ws_mensal = write_data_sheet(wb, mensal, "dados_mensal")
    ws_atraso = write_data_sheet(wb, atraso, "dados_atraso_nota")
    ws_simul = write_data_sheet(wb, simul, "dados_simulacao")
    ws_cat = write_data_sheet(wb, categoria, "dados_categoria")
    ws_est = write_data_sheet(wb, estado, "dados_estado")

    # ---- canvas do dashboard (fundo cinza claro atras de tudo) ----
    canvas_fill = PatternFill("solid", fgColor=CINZA_CLARO)
    for r in range(1, 45):
        for col in range(1, 19):
            dash.cell(row=r, column=col).fill = canvas_fill

    # largura base das colunas
    for col in range(1, 19):
        dash.column_dimensions[get_column_letter(col)].width = 9

    # ---- faixa de titulo (header) ----
    dash.merge_cells("A1:R2")
    header_fill = PatternFill("solid", fgColor=AZUL)
    for r in (1, 2):
        for col in range(1, 19):
            dash.cell(row=r, column=col).fill = header_fill
    h = dash["A1"]
    h.value = "OLIST    O paradoxo do crescimento: vende mais, mas a entrega ameaça o futuro"
    h.font = Font(bold=True, size=18, color=BRANCO)
    h.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    dash.row_dimensions[1].height = 24
    dash.row_dimensions[2].height = 16

    dash.merge_cells("A3:R3")
    s = dash["A3"]
    s.value = "Tech Challenge · Data Analytics · Olist 2017–2018 · Público: investidores e diretoria"
    s.font = Font(italic=True, size=10, color=CINZA)
    s.alignment = Alignment(horizontal="left", indent=1)

    # ---- linha de KPIs (5 cartoes) ----
    kv = {row["indicador"]: row for _, row in kpis.iterrows()}
    cards = [
        ("RECEITA TOTAL", "R$ 13,5 mi", AZUL),
        ("PEDIDOS", "99.092", AZUL),
        ("TICKET MEDIO", "R$ 137,68", AZUL),
        ("TAXA DE RECOMPRA", "3,1%", VERMELHO),
        ("PEDIDOS ATRASADOS", "8,1%", VERMELHO),
    ]
    start_row = 4
    for i, (titulo, valor, cor) in enumerate(cards):
        kpi_card(dash, anchor_col=1 + i * 3, row=start_row, titulo=titulo, valor=valor, cor=cor)

    # banda de contexto sob os KPIs
    dash.merge_cells("A7:R7")
    band = dash["A7"]
    band.value = ("Negócio movido a aquisição (recompra 10x abaixo do mercado ~28%)  →  reputação é o ativo crítico  "
                  "→  atraso derruba a nota (4,29 → 2,57) e expõe R$ 1,2 mi de receita.")
    band.font = Font(size=9, italic=True, color=CINZA)
    band.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    # ============ GRAFICO A: Receita mensal ============
    n = len(mensal)
    bar = BarChart()
    bar.type = "col"
    bar.title = "Crescimento da receita ao longo do tempo (estabiliza em 2018)"
    bar.height = 7.0
    bar.width = 13.5
    data = Reference(ws_mensal, min_col=3, min_row=1, max_row=n + 1)  # receita
    cats = Reference(ws_mensal, min_col=1, min_row=2, max_row=n + 1)  # ano_mes
    bar.add_data(data, titles_from_data=True)
    bar.set_categories(cats)
    bar.series[0].graphicalProperties.solidFill = AZUL
    bar.legend = None
    style_chart(bar)
    dash.add_chart(bar, "A9")

    # ============ GRAFICO B: Atraso x Nota (o "aha") ============
    b = BarChart()
    b.type = "col"
    b.title = "Quanto mais atrasa, pior a nota (dose-resposta)"
    b.height = 7.0
    b.width = 13.5
    bdata = Reference(ws_atraso, min_col=2, min_row=1, max_row=len(atraso) + 1)  # nota_media
    bcats = Reference(ws_atraso, min_col=1, min_row=2, max_row=len(atraso) + 1)
    b.add_data(bdata, titles_from_data=True)
    b.set_categories(bcats)
    b.series[0].graphicalProperties.solidFill = VERMELHO
    b.dataLabels = value_labels()
    b.legend = None
    style_chart(b)
    dash.add_chart(b, "J9")

    # ============ GRAFICO C: Simulacao (valor de agir) ============
    c = BarChart()
    c.type = "col"
    c.title = "Valor de agir: receita protegida ao reduzir atrasos (R$)"
    c.height = 7.0
    c.width = 13.5
    cdata = Reference(ws_simul, min_col=3, min_row=1, max_row=len(simul) + 1)  # receita_protegida
    ccats = Reference(ws_simul, min_col=1, min_row=2, max_row=len(simul) + 1)
    c.add_data(cdata, titles_from_data=True)
    c.set_categories(ccats)
    c.series[0].graphicalProperties.solidFill = VERDE
    c.dataLabels = value_labels()
    c.legend = None
    style_chart(c)
    dash.add_chart(c, "A27")

    # ============ GRAFICO D: Top categorias por receita ============
    d = BarChart()
    d.type = "bar"
    d.title = "Top categorias por receita (R$)"
    d.height = 7.0
    d.width = 13.5
    ddata = Reference(ws_cat, min_col=3, min_row=1, max_row=len(categoria) + 1)  # receita
    dcats = Reference(ws_cat, min_col=1, min_row=2, max_row=len(categoria) + 1)
    d.add_data(ddata, titles_from_data=True)
    d.set_categories(dcats)
    d.series[0].graphicalProperties.solidFill = AZUL
    d.legend = None
    style_chart(d)
    dash.add_chart(d, "J27")

    # rodape sintese
    dash.merge_cells("A43:R43")
    foot = dash["A43"]
    foot.fill = PatternFill("solid", fgColor=AZUL)
    foot.value = ("Logística não é custo: é a alavanca que protege o crescimento.   "
                  "Reduzir atrasos em 50% protege ~R$ 579 mil e evita ~1.755 detratores.")
    foot.font = Font(bold=True, size=11, color=BRANCO)
    foot.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    for col in range(1, 19):
        dash.cell(row=43, column=col).fill = PatternFill("solid", fgColor=AZUL)
    dash.row_dimensions[43].height = 22

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"OK -> {OUT.relative_to(ROOT)}")
    print("Abas:", wb.sheetnames)


if __name__ == "__main__":
    main()
